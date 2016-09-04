import sys, os
import openpyxl as pyxl
import datetime
import pandas as pd

def parse_excel(fullfile):
	wb = pyxl.load_workbook(fullfile, read_only = True,
		data_only = True)
	data_dict = {
		'year': [],
		'month' : [],
		'state_code' : [],
		'producer_type' : [],
		'energy_source' : [],
		'total_energy' : []
	}
	worksheets = wb.worksheets
	for worksheet in worksheets:
		ws = wb.get_sheet_by_name(worksheet.title)
		print('Parsing Sheet: ' + worksheet.title)

		num_rows = ws.max_row
		for row in ws.iter_rows('A2:F'+str(num_rows)):
			data_dict['year'].append(int(row[0].value))
			data_dict['month'].append(int(row[1].value))
			data_dict['state_code'].append(str(row[2].value))
			data_dict['producer_type'].append(str(row[3].value))
			data_dict['energy_source'].append(str(row[4].value))
			data_dict['total_energy'].append(int(row[5].value))
		df = pd.DataFrame(data_dict)
	return df


if __name__ == '__main__':
	filedir = '/home/matt/ISE5144_project/data/raw/'
	destdir = '/home/matt/ISE5144_project/data/cleaned/'
	files = [
		'nrel_energy_generation.xlsx',
		'nrel_energy_consumption.xlsx'
	]
	filename = 'nrel_energy_generation.xlsx'
	for file in files:
		fullfile = filedir+file
		destfile = file.split('.')[0]+'.csv'
		print('Parsing File: ' + file)
		df = parse_excel(fullfile)
		df.to_csv(destdir + destfile)
	print('done ... yay!')